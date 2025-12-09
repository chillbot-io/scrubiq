"""Excel spreadsheet (.xlsx) extraction."""

from pathlib import Path

from .base import Extractor, ExtractionError

try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class XlsxExtractor(Extractor):
    """Extract text from Excel spreadsheets."""

    @property
    def extensions(self) -> list[str]:
        return [".xlsx", ".xlsm"]

    def extract(self, path: Path) -> str:
        """Extract text from all cells in all sheets."""
        if not HAS_OPENPYXL:
            raise ExtractionError("openpyxl not installed. Run: pip install openpyxl")

        try:
            # data_only=True to get computed values, read_only=True for performance
            wb = load_workbook(path, data_only=True, read_only=True)
            text_parts = []

            for sheet in wb.worksheets:
                # Add sheet name as context
                text_parts.append(f"[Sheet: {sheet.title}]")

                for row in sheet.iter_rows():
                    row_text = []
                    for cell in row:
                        if cell.value is not None:
                            row_text.append(str(cell.value))
                    if row_text:
                        text_parts.append(" | ".join(row_text))

            wb.close()
            return "\n".join(text_parts)

        except Exception as e:
            raise ExtractionError(f"Failed to extract from {path}: {e}")
